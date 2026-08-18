"""
clean_and_export.py
-------------------
Cleans the raw checkpoint data and exports the final dataset.

Final schema (exactly, in order):
  1. full_name
  2. specialization
  3. qualification
  4. experience
  5. location
  6. registration_details
  7. practo_profile_url

Rules enforced:
  - Missing values → empty string (never dropped)
  - No field merging — columns are strictly separate
  - experience stored as plain text ("12 years"), no reformatting
  - registration_details = full registration text as scraped
  - practo_profile_url = exact sitemap URL, unmodified
  - Deduplication on practo_profile_url (keep first occurrence)
  - Exports to both .csv and formatted .xlsx
"""

import sys
import os
import logging
from pathlib import Path

# Fix Unicode output on Windows consoles (cp1252 → utf-8)
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import pandas as pd
from dotenv import load_dotenv

# Load .env for Neon DB credentials
load_dotenv(dotenv_path=Path(__file__).parent.parent / ".env")

# ── Paths ──────────────────────────────────────────────────────────────────
BASE_DIR       = Path(__file__).parent.parent
CHECKPOINT_CSV = BASE_DIR / "output" / "practo_doctors_checkpoint.csv"
FINAL_CSV      = BASE_DIR / "output" / "practo_doctors_final.csv"
FINAL_XLSX     = BASE_DIR / "output" / "practo_doctors_final.xlsx"
LOG_DIR        = BASE_DIR / "logs"

# ── Exact final column order ───────────────────────────────────────────────
FINAL_COLUMNS = [
    "full_name",
    "first_name",
    "last_name",
    "specialization",
    "specialization_alias",
    "qualification",
    "experience",
    "address_of_clinic",
    "location",
    "registration_details",
    "practo_profile_url",
]

# Legacy column names from older checkpoint files → map to new names
COLUMN_RENAMES = {
    "profile_url":  "practo_profile_url",
    "registration": "registration_details",
}

# ── Logging ────────────────────────────────────────────────────────────────
LOG_DIR.mkdir(exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(LOG_DIR / "clean_export.log", encoding="utf-8"),
    ],
)
log = logging.getLogger("clean_export")


# ── Helpers ────────────────────────────────────────────────────────────────

def _rename_legacy_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Handle older checkpoint files that used different column names."""
    return df.rename(columns={k: v for k, v in COLUMN_RENAMES.items() if k in df.columns})


def _enforce_schema(df: pd.DataFrame) -> pd.DataFrame:
    """
    Ensure the DataFrame has exactly FINAL_COLUMNS in the right order.
    - Adds any missing column as empty string
    - Drops any extra columns not in FINAL_COLUMNS
    - Reorders to FINAL_COLUMNS
    """
    for col in FINAL_COLUMNS:
        if col not in df.columns:
            log.warning(f"  Column '{col}' not found in data — adding as empty")
            df[col] = ""
    return df[FINAL_COLUMNS]


def _clean_values(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normalise values:
    - Strip leading/trailing whitespace from all string columns
    - Replace placeholder strings and NaN/None with empty string
      (rule: missing fields must be blank, NOT dropped)
    - Leave experience and practo_profile_url untouched
    """
    placeholders = {"None", "none", "N/A", "n/a", "NA", "na", "nan", "-", "null"}

    for col in df.columns:
        # Convert to string first so we can do string ops uniformly
        df[col] = df[col].astype(str).str.strip()
        # Treat Python/pandas string representations of nulls as empty
        df[col] = df[col].replace({p: "" for p in placeholders | {"nan", "NaT", "<NA>"}})

    return df


def _null_report(df: pd.DataFrame) -> None:
    """Print a clear per-column completeness report."""
    total = len(df)
    print()
    print("-" * 68)
    print(f"  DATA COMPLETENESS REPORT  (total rows: {total:,})")
    print("-" * 68)
    print(f"  {'#':<3}  {'Column':<25}  {'Empty':>7}  {'%':>7}  {'Filled':>8}")
    print("  " + "-" * 62)
    for i, col in enumerate(FINAL_COLUMNS, 1):
        empty  = int((df[col] == "").sum())
        filled = total - empty
        pct    = (empty / total * 100) if total > 0 else 0.0
        flag   = "  !" if pct > 20 else ""
        print(f"  {i:<3}  {col:<25}  {empty:>7,}  {pct:>6.1f}%  {filled:>8,}{flag}")
    print("-" * 68)
    print()


def _export_xlsx(df: pd.DataFrame, path: Path) -> None:
    """Write a nicely formatted .xlsx file."""
    try:
        import openpyxl  # noqa: F401
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Practo Doctors Pune")
            ws = writer.sheets["Practo Doctors Pune"]

            # ── Header row styling ─────────────────────────────────────────
            header_fill = PatternFill("solid", fgColor="1E2D78")
            header_font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
            center      = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for col_idx, cell in enumerate(ws[1], start=1):
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = center
                # Auto-width
                col_letter = get_column_letter(col_idx)
                header_len = len(str(cell.value) or "")
                data_len   = max(
                    (len(str(ws.cell(row=r, column=col_idx).value or ""))
                     for r in range(2, min(ws.max_row + 1, 300))),
                    default=0,
                )
                ws.column_dimensions[col_letter].width = min(max(header_len, data_len) + 3, 60)

            # ── Data row styling (alternating bands) ──────────────────────
            thin = Border(bottom=Side(style="hair", color="DDDDDD"))
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                fill_color = "EEF0FF" if row_idx % 2 == 0 else "FFFFFF"
                row_fill   = PatternFill("solid", fgColor=fill_color)
                for cell in row:
                    cell.fill      = row_fill
                    cell.border    = thin
                    cell.font      = Font(size=10, name="Calibri")
                    cell.alignment = Alignment(vertical="center", wrap_text=False)

            # Freeze header
            ws.freeze_panes = "A2"
            # Auto-filter
            ws.auto_filter.ref = ws.dimensions

        log.info(f"  XLSX written → {path}")

    except ImportError:
        log.warning("openpyxl not available. Writing plain xlsx...")
        df.to_excel(path, index=False)
        log.info(f"  XLSX written (plain) → {path}")


import argparse

# ── Main ───────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Clean and export Practo doctor data by location.")
    parser.add_argument(
        "-l", "--location",
        type=str,
        default="pune",
        help="Target location keyword (e.g. pune, mumbai, delhi, margao). Default: pune"
    )
    parser.add_argument(
        "-i", "--input",
        type=str,
        default=None,
        help="Path to checkpoint CSV."
    )
    parser.add_argument(
        "--csv",
        type=str,
        default=None,
        help="Output CSV path."
    )
    parser.add_argument(
        "--xlsx",
        type=str,
        default=None,
        help="Output XLSX path."
    )

    args = parser.parse_args()
    loc = args.location.strip().lower()

    if args.input:
        checkpoint_csv = Path(args.input)
    else:
        checkpoint_csv = BASE_DIR / "output" / ("practo_doctors_checkpoint.csv" if loc == "pune" else f"practo_doctors_checkpoint_{loc}.csv")

    if args.csv:
        final_csv = Path(args.csv)
    else:
        final_csv = BASE_DIR / "output" / ("practo_doctors_final.csv" if loc == "pune" else f"practo_doctors_final_{loc}.csv")

    if args.xlsx:
        final_xlsx = Path(args.xlsx)
    else:
        final_xlsx = BASE_DIR / "output" / ("practo_doctors_final.xlsx" if loc == "pune" else f"practo_doctors_final_{loc}.xlsx")

    # 1. Load checkpoint
    if not checkpoint_csv.exists():
        log.error(f"Checkpoint file not found: {checkpoint_csv}")
        sys.exit(1)

    log.info(f"Loading: {checkpoint_csv}")
    df = pd.read_csv(checkpoint_csv, dtype=str)
    log.info(f"  Raw rows loaded: {len(df):,}  |  Columns: {list(df.columns)}")

    # 2. Handle legacy column names
    df = _rename_legacy_columns(df)

    # 3. Enforce exact final schema (add missing cols, drop extras, reorder)
    df = _enforce_schema(df)

    # 4. Clean values — whitespace, placeholder strings → empty string
    df = _clean_values(df)

    # 5. Deduplicate on practo_profile_url
    before = len(df)
    df.drop_duplicates(subset=["practo_profile_url"], keep="first", inplace=True)
    dupes = before - len(df)
    log.info(f"  Duplicates removed (by practo_profile_url): {dupes:,}")

    # 6. Drop rows where practo_profile_url is blank (data integrity guard)
    before = len(df)
    df = df[df["practo_profile_url"] != ""]
    dropped_empty_url = before - len(df)
    if dropped_empty_url:
        log.warning(f"  Dropped {dropped_empty_url} rows with blank practo_profile_url")

    df.reset_index(drop=True, inplace=True)
    log.info(f"  Final row count: {len(df):,}")

    # 7. Null / completeness report (printed to stdout)
    _null_report(df)

    # 8. Export CSV
    df.to_csv(final_csv, index=False, encoding="utf-8-sig")  # utf-8-sig for Excel compat
    log.info(f"  CSV written  → {final_csv}")

    # 9. Export XLSX
    _export_xlsx(df, final_xlsx)

    # 10. Upload to Neon DB (if configured)
    db_url = os.getenv("NEON_DATABASE_URL", "")
    if db_url and "your_user" not in db_url:
        log.info("NEON_DATABASE_URL found — uploading to Neon DB...")
        try:
            # Import here to avoid hard dependency when Neon is not used
            from neon_upload import _connect, _upload
            conn = _connect()
            try:
                uploaded = _upload(df, loc, conn)
                log.info(f"  ✓ Neon DB upload complete — {uploaded:,} rows upserted.")
            except Exception as e:
                conn.rollback()
                log.error(f"  Neon DB upload failed: {e}")
            finally:
                conn.close()
        except ImportError:
            log.warning("  psycopg2 not installed. Run: pip install psycopg2-binary")
    else:
        log.info("  Neon DB upload skipped (NEON_DATABASE_URL not configured in .env).")

    # 11. Final summary
    print("=" * 68)
    print("EXPORT COMPLETE")
    print(f"  Total rows in final dataset : {len(df):,}")
    print(f"  Duplicates removed          : {dupes:,}")
    print(f"  Output CSV                  : {final_csv}")
    print(f"  Output XLSX                 : {final_xlsx}")
    if db_url and "your_user" not in db_url:
        print(f"  Neon DB table              : practo_doctors (tag='{loc}')")
    print("=" * 68)


if __name__ == "__main__":
    main()
